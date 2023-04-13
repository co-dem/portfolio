import openpyxl


book = openpyxl.open('C:\\Users\\mamed\\Desktop\\trump_bot\\product_list.xlsx')
sheet = book.active
stats = {}

def update_database() -> any:
    global book, sheet

    book.close()

    book = openpyxl.open('product_list.xlsx')
    sheet = book.active

    return sheet

def get_products(xl = sheet) -> list:
    tastes = []

    for i in range(1, 300):
        if xl.cell(row = i, column = 1).value != None:
            tastes.append(xl.cell(row = i, column = 1).value.lower())
        else:
            break
    print(f'list updated\n{tastes}')
    
    return tastes

def check_product(name) -> bool:
    global sheet

    if name.lower() in get_products(xl = sheet):
        return True

    else:
        return False

def upload_products() -> dict:
    global stats

    for i in get_products():
        prod_exists = stats.get(i)
        if prod_exists == None:
            stats[i] = 0
    
    return stats

def changestats(product_name: str):
    global stats
    
    prod_check = stats.get(product_name.lower())
    if prod_check >= 0:
        stats[product_name.lower()] += 1
        return stats
    else:
        return ValueError

def clear_stats() -> dict:
    global stats

    for i in stats:
        stats[i] = 0
    return stats

def showstats() -> dict:
    global stats
    return stats

def get_price(product_name) -> int:
    global sheet

    for i in range(300):
        if sheet.cell(row = i+1, column = 1).value == product_name.lower():
            return sheet.cell(row = i+1, column = 2).value
        
def get_sale(product_name) -> int:
    global sheet

    for i in range(300):
        if sheet.cell(row = i+1, column = 1).value == product_name.lower():
            if bool(sheet.cell(row = i+1, column = 3).value) != False:
                return sheet.cell(row = i+1, column = 3).value
    return 100
        

#| coded by codem